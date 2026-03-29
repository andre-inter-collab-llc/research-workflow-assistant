"""Shared SQLite storage for MCP search results.

Provides per-project result persistence so searches can be queried
later from R (RSQLite/DBI) or Python (sqlite3 stdlib).

Database location: {project_path}/data/search_results.db
"""

import csv
import json
import logging
import re
import sqlite3
import subprocess
import sys
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

_SCHEMA = """\
CREATE TABLE IF NOT EXISTS searches (
    search_id INTEGER PRIMARY KEY AUTOINCREMENT,
    source TEXT NOT NULL,
    query TEXT NOT NULL,
    timestamp TEXT NOT NULL,
    total_count INTEGER,
    parameters_json TEXT
);

CREATE TABLE IF NOT EXISTS results (
    result_id INTEGER PRIMARY KEY AUTOINCREMENT,
    search_id INTEGER NOT NULL REFERENCES searches(search_id),
    source TEXT NOT NULL,
    doi TEXT,
    pmid TEXT,
    title TEXT,
    authors_json TEXT,
    journal TEXT,
    year TEXT,
    volume TEXT,
    issue TEXT,
    pages TEXT,
    abstract TEXT,
    extra_json TEXT,
    retrieved_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_results_doi
    ON results(doi) WHERE doi IS NOT NULL AND doi != '';
CREATE INDEX IF NOT EXISTS idx_results_pmid
    ON results(pmid) WHERE pmid IS NOT NULL AND pmid != '';
CREATE INDEX IF NOT EXISTS idx_results_search_id
    ON results(search_id);

CREATE TABLE IF NOT EXISTS attachments (
    attachment_id INTEGER PRIMARY KEY AUTOINCREMENT,
    result_id INTEGER NOT NULL REFERENCES results(result_id),
    file_path TEXT NOT NULL,
    mime_type TEXT,
    label TEXT,
    added_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_attachments_result
    ON attachments(result_id);

CREATE TABLE IF NOT EXISTS notes (
    note_id INTEGER PRIMARY KEY AUTOINCREMENT,
    result_id INTEGER NOT NULL REFERENCES results(result_id),
    note_type TEXT NOT NULL DEFAULT 'general',
    content TEXT NOT NULL,
    tags_json TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_notes_result
    ON notes(result_id);

CREATE TABLE IF NOT EXISTS annotations (
    annotation_id INTEGER PRIMARY KEY AUTOINCREMENT,
    attachment_id INTEGER NOT NULL REFERENCES attachments(attachment_id),
    result_id INTEGER NOT NULL REFERENCES results(result_id),
    page INTEGER,
    annotation_type TEXT,
    color TEXT,
    content TEXT,
    comment TEXT,
    extracted_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_annotations_attachment
    ON annotations(attachment_id);
CREATE INDEX IF NOT EXISTS idx_annotations_result
    ON annotations(result_id);
"""

_DEDUP_VIEW = """\
CREATE VIEW IF NOT EXISTS deduplicated_results AS
SELECT
    MIN(result_id) AS result_id,
    doi,
    MAX(pmid) AS pmid,
    MAX(title) AS title,
    MAX(authors_json) AS authors_json,
    MAX(journal) AS journal,
    MAX(year) AS year,
    GROUP_CONCAT(DISTINCT source) AS sources,
    COUNT(*) AS occurrence_count,
    MIN(retrieved_at) AS first_retrieved
FROM results
GROUP BY
    CASE
        WHEN doi IS NOT NULL AND doi != '' THEN 'doi:' || doi
        WHEN pmid IS NOT NULL AND pmid != '' THEN 'pmid:' || pmid
        ELSE 'id:' || result_id
    END;
"""

# Fields extracted into dedicated columns (everything else goes to extra_json)
_CORE_KEYS = frozenset(
    {
        "doi",
        "pmid",
        "title",
        "authors",
        "journal",
        "year",
        "volume",
        "issue",
        "pages",
        "page",
        "abstract",
        "source",
        "fulljournalname",
        "venue",
        "publication_year",
        "pubdate",
    }
)


def _db_path(project_path: str) -> Path:
    """Return the SQLite database path for a project."""
    return Path(project_path) / "data" / "search_results.db"


def init_db(project_path: str) -> sqlite3.Connection:
    """Initialize the search results database for a project.

    Creates the data/ directory and search_results.db if they don't exist.
    Returns an open connection.
    """
    db = _db_path(project_path)
    db.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db))
    conn.executescript(_SCHEMA)
    # Create dedup view (separate because CREATE VIEW IF NOT EXISTS
    # doesn't play well inside executescript with other statements on some builds)
    try:
        conn.executescript(_DEDUP_VIEW)
    except sqlite3.OperationalError:
        pass  # View already exists
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA foreign_keys=ON")
    return conn


def _normalize_authors(authors: Any) -> str:
    """Normalize authors from various formats to a JSON array of strings."""
    if isinstance(authors, str):
        # Europe PMC returns comma-separated string
        return json.dumps([a.strip() for a in authors.split(",") if a.strip()])
    if isinstance(authors, list):
        if authors and isinstance(authors[0], dict):
            # Semantic Scholar returns [{name, author_id}, ...]
            return json.dumps([a.get("name", str(a)) for a in authors])
        return json.dumps(authors)
    return json.dumps([])


def _extract_year(result: dict[str, Any]) -> str:
    """Extract a 4-digit year from various result formats."""
    for key in ("year", "publication_year", "pubdate"):
        val = result.get(key)
        if val is not None:
            s = str(val).strip()
            if len(s) >= 4 and s[:4].isdigit():
                return s[:4]
            if s:
                return s
    return ""


def store_results(
    project_path: str,
    source: str,
    query: str,
    results: list[dict[str, Any]],
    total_count: int | None = None,
    parameters: dict[str, Any] | None = None,
) -> int:
    """Store search results in the project database.

    Args:
        project_path: Absolute path to the project directory.
        source: Search source identifier (e.g., 'pubmed', 'openalex').
        query: The search query that was executed.
        results: List of result dicts from the search server's format function.
        total_count: Total results available (not just returned).
        parameters: Search parameters for reproducibility.

    Returns:
        The search_id of the stored search.
    """
    conn = init_db(project_path)
    try:
        now = datetime.now(UTC).isoformat()

        cursor = conn.execute(
            "INSERT INTO searches (source, query, timestamp, total_count, parameters_json) "
            "VALUES (?, ?, ?, ?, ?)",
            (source, query, now, total_count, json.dumps(parameters) if parameters else None),
        )
        search_id = cursor.lastrowid

        for r in results:
            authors_json = _normalize_authors(r.get("authors", []))
            journal = (
                r.get("journal", "") or r.get("fulljournalname", "") or r.get("venue", "") or ""
            )
            year = _extract_year(r)
            pages = r.get("pages", "") or r.get("page", "") or ""
            extra = {k: v for k, v in r.items() if k not in _CORE_KEYS and v}

            conn.execute(
                "INSERT INTO results "
                "(search_id, source, doi, pmid, title, authors_json, journal, "
                "year, volume, issue, pages, abstract, extra_json, retrieved_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    search_id,
                    source,
                    r.get("doi", "") or "",
                    r.get("pmid", "") or "",
                    r.get("title", "") or "",
                    authors_json,
                    journal,
                    year,
                    r.get("volume", "") or "",
                    r.get("issue", "") or "",
                    pages,
                    r.get("abstract", "") or "",
                    json.dumps(extra) if extra else None,
                    now,
                ),
            )

        conn.commit()
        return search_id
    finally:
        conn.close()


def get_results(
    project_path: str,
    source: str | None = None,
    query: str | None = None,
    deduplicated: bool = False,
) -> list[dict[str, Any]]:
    """Query stored results from the project database.

    Args:
        project_path: Absolute path to the project directory.
        source: Optional filter by source (e.g., 'pubmed').
        query: Optional filter by query substring.
        deduplicated: If True, return deduplicated results across sources.

    Returns:
        List of result dicts.
    """
    db = _db_path(project_path)
    if not db.exists():
        return []

    conn = init_db(project_path)
    conn.row_factory = sqlite3.Row
    try:
        table = "deduplicated_results" if deduplicated else "results"
        where_clauses: list[str] = []
        params: list[str] = []

        if source and not deduplicated:
            where_clauses.append("source = ?")
            params.append(source)
        if query and not deduplicated:
            where_clauses.append("search_id IN (SELECT search_id FROM searches WHERE query LIKE ?)")
            params.append(f"%{query}%")

        sql = f"SELECT * FROM {table}"
        if where_clauses:
            sql += " WHERE " + " AND ".join(where_clauses)

        rows = conn.execute(sql, params).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def get_searches(project_path: str) -> list[dict[str, Any]]:
    """List all recorded searches for a project."""
    db = _db_path(project_path)
    if not db.exists():
        return []

    conn = init_db(project_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT s.*, COUNT(r.result_id) AS result_count "
            "FROM searches s LEFT JOIN results r ON s.search_id = r.search_id "
            "GROUP BY s.search_id ORDER BY s.timestamp DESC"
        ).fetchall()
        return [dict(row) for row in rows]
    finally:
        conn.close()


def export_results_csv(
    project_path: str,
    output_path: str | None = None,
    deduplicated: bool = False,
) -> str:
    """Export stored results to CSV. Returns the output file path."""
    results = get_results(project_path, deduplicated=deduplicated)
    if not results:
        return ""

    if output_path is None:
        suffix = "_deduplicated" if deduplicated else ""
        output_path = str(Path(project_path) / "data" / f"search_results{suffix}.csv")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    fieldnames = list(results[0].keys())
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)

    return output_path


def export_results_excel(
    project_path: str,
    output_path: str | None = None,
    deduplicated: bool = False,
) -> str:
    """Export stored results to an Excel (.xlsx) workbook.

    Creates two sheets:
    - **All Results** with every result row
    - **Searches** with summary metadata for each search

    Returns the output file path, or empty string if no results.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    results = get_results(project_path, deduplicated=deduplicated)
    if not results:
        return ""

    if output_path is None:
        suffix = "_deduplicated" if deduplicated else ""
        output_path = str(Path(project_path) / "data" / f"search_results{suffix}.xlsx")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    link_font = Font(color="0563C1", underline="single")

    def _doi_url(doi: str) -> str:
        doi = doi.strip()
        if doi.startswith("http"):
            return doi
        return f"https://doi.org/{doi}"

    def _pmid_url(pmid: str) -> str:
        pmid = pmid.strip()
        if pmid.startswith("http"):
            return pmid
        return f"https://pubmed.ncbi.nlm.nih.gov/{pmid}/"

    wb = Workbook()

    # --- Results sheet ---
    ws = wb.active
    ws.title = "All Results"

    # Prioritised column order for human review.  Any remaining columns
    # from the result dicts are appended afterwards in their original order.
    priority_cols = [
        "title",
        "authors_json",
        "journal",
        "year",
        "doi",
        "pmid",
        "abstract",
        "source",
    ]
    header_labels = {"authors_json": "Authors"}

    raw_keys = list(results[0].keys())
    ordered: list[str] = [c for c in priority_cols if c in raw_keys]
    ordered += [k for k in raw_keys if k not in ordered]

    display_headers = [header_labels.get(h, h) for h in ordered]
    ws.append(display_headers)

    doi_col = ordered.index("doi") + 1 if "doi" in ordered else None
    pmid_col = ordered.index("pmid") + 1 if "pmid" in ordered else None

    for row in results:
        ws.append([row.get(h, "") for h in ordered])

    # Add hyperlinks for DOI and PMID columns
    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if doi_col:
            cell = row_cells[doi_col - 1]
            val = str(cell.value or "").strip()
            if val:
                cell.hyperlink = _doi_url(val)
                cell.font = link_font
        if pmid_col:
            cell = row_cells[pmid_col - 1]
            val = str(cell.value or "").strip()
            if val:
                cell.hyperlink = _pmid_url(val)
                cell.font = link_font

    # Auto-width (capped at 60; abstract uses text-wrap instead)
    for col_idx, h in enumerate(ordered, 1):
        if h == "abstract":
            ws.column_dimensions[get_column_letter(col_idx)].width = 80
            for row_cells in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for cell in row_cells:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
            continue
        max_len = len(str(display_headers[col_idx - 1]))
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                val_len = len(str(cell.value or ""))
                if val_len > max_len:
                    max_len = val_len
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    # Format as Excel Table with auto-filter
    end_col = get_column_letter(len(ordered))
    results_table = Table(
        displayName="Results",
        ref=f"A1:{end_col}{ws.max_row}",
    )
    results_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(results_table)

    # --- Searches sheet ---
    searches = get_searches(project_path)
    if searches:
        ws2 = wb.create_sheet("Searches")
        s_headers = list(searches[0].keys())
        ws2.append(s_headers)
        for s in searches:
            ws2.append([s.get(h, "") for h in s_headers])
        for col_idx, h in enumerate(s_headers, 1):
            ws2.column_dimensions[get_column_letter(col_idx)].width = min(
                max(len(str(h)), 12) + 2, 40
            )
        s_end_col = get_column_letter(len(s_headers))
        searches_table = Table(
            displayName="Searches",
            ref=f"A1:{s_end_col}{ws2.max_row}",
        )
        searches_table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws2.add_table(searches_table)

    wb.save(output_path)
    return output_path


# ---------------------------------------------------------------------------
# Script-first search execution
# ---------------------------------------------------------------------------

_TEMPLATE_FUNCS: dict[str, str] = {
    "pubmed": "pubmed_script",
    "openalex": "openalex_script",
    "semantic_scholar": "semantic_scholar_script",
    "europe_pmc": "europe_pmc_script",
    "crossref": "crossref_script",
}


def _get_template_func(source: str):  # noqa: ANN202
    """Lazily import and return the script template function for *source*."""
    from rwa_result_store.script_templates import (  # lazy import
        crossref_script,
        europe_pmc_script,
        openalex_script,
        pubmed_script,
        semantic_scholar_script,
    )

    template_map = {
        "pubmed": pubmed_script,
        "openalex": openalex_script,
        "semantic_scholar": semantic_scholar_script,
        "europe_pmc": europe_pmc_script,
        "crossref": crossref_script,
    }
    return template_map.get(source)


def generate_search_script(
    project_path: str,
    source: str,
    query: str,
    parameters: dict[str, Any],
) -> tuple[str, str] | None:
    """Generate a standalone search script **without** executing it.

    Use this for the *draft-then-approve* workflow: the agent presents
    the script to the user for review before calling
    :func:`execute_search_script`.

    Args:
        project_path: Absolute path to the project directory.
        source: Search source identifier (e.g., ``"pubmed"``).
        query: The search query string.
        parameters: Dict of search parameters (forwarded to the template).

    Returns:
        ``(script_path, script_content)`` on success, or ``None`` if
        template generation fails.
    """
    func = _get_template_func(source)
    if func is None:
        logger.warning("No script template for source: %s", source)
        return None

    try:
        script_content = func(query=query, project_path=project_path, **parameters)
    except Exception:
        logger.warning("Failed to generate script for %s", source, exc_info=True)
        return None

    ts = datetime.now(UTC).strftime("%Y%m%d_%H%M%S")
    scripts_dir = Path(project_path) / "scripts"
    scripts_dir.mkdir(parents=True, exist_ok=True)
    script_path = scripts_dir / f"search_{source}_{ts}.py"
    script_path.write_text(script_content, encoding="utf-8")

    return (str(script_path), script_content)


def execute_search_script(
    project_path: str,
    script_path: str,
) -> tuple[list[dict[str, Any]], int, int] | None:
    """Execute a previously generated search script and return results.

    The script is expected to print a ``search_id`` integer to stdout
    (or ``NO_RESULTS``). Results are read back from the project's
    SQLite database.

    Args:
        project_path: Absolute path to the project directory.
        script_path: Absolute path to the search script to run.

    Returns:
        ``(results_list, total_count, search_id)`` on success,
        or ``None`` if execution fails.
    """
    script_path_obj = Path(script_path)
    if not script_path_obj.exists():
        logger.warning("Script not found: %s", script_path)
        return None

    try:
        result = subprocess.run(
            [sys.executable, str(script_path_obj)],
            capture_output=True,
            text=True,
            timeout=120,
        )
    except (subprocess.TimeoutExpired, OSError) as exc:
        logger.warning("Script execution failed: %s", exc)
        return None

    if result.returncode != 0:
        logger.warning(
            "Script %s exited with code %d. stderr: %s",
            script_path_obj.name,
            result.returncode,
            result.stderr[:500],
        )
        return None

    stdout = result.stdout.strip()
    if not stdout or stdout == "NO_RESULTS":
        return ([], 0, 0)

    try:
        search_id = int(stdout)
    except ValueError:
        logger.warning("Could not parse search_id from script output: %r", stdout)
        return None

    db = _db_path(project_path)
    if not db.exists():
        return None

    conn = init_db(project_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute("SELECT * FROM results WHERE search_id = ?", (search_id,)).fetchall()
        results = [dict(row) for row in rows]

        row = conn.execute(
            "SELECT total_count FROM searches WHERE search_id = ?", (search_id,)
        ).fetchone()
        total_count = row["total_count"] if row else 0
    finally:
        conn.close()

    return (results, total_count, search_id)


def generate_and_run_script(
    project_path: str,
    source: str,
    query: str,
    parameters: dict[str, Any],
) -> tuple[list[dict[str, Any]], int, int, str] | None:
    """Generate a standalone search script, execute it, and read results.

    This is a convenience wrapper that calls :func:`generate_search_script`
    followed by :func:`execute_search_script`.

    Args:
        project_path: Absolute path to the project directory.
        source: Search source identifier (e.g., ``"pubmed"``).
        query: The search query string.
        parameters: Dict of search parameters (passed to the template function).

    Returns:
        ``(results_list, total_count, search_id, script_path)`` on success,
        or ``None`` if script generation/execution fails.
    """
    gen_result = generate_search_script(project_path, source, query, parameters)
    if gen_result is None:
        return None

    script_path, _script_content = gen_result

    exec_result = execute_search_script(project_path, script_path)
    if exec_result is None:
        return None

    results, total_count, search_id = exec_result
    return (results, total_count, search_id, script_path)


# ---------------------------------------------------------------------------
# Bibliographic format exports (BibTeX, RIS, CSL-JSON)
# ---------------------------------------------------------------------------


def _parse_authors(authors_json: str | None) -> list[str]:
    """Parse authors_json column into a list of name strings."""
    if not authors_json:
        return []
    try:
        parsed = json.loads(authors_json)
        if isinstance(parsed, list):
            return [str(a) for a in parsed]
    except (json.JSONDecodeError, TypeError):
        pass
    return []


def _make_citekey(authors: list[str], year: str, doi: str, result_id: int) -> str:
    """Generate a BibTeX-safe cite key like ``AuthorYYYY``."""
    if authors:
        first = authors[0]
        # Take the last name (before any comma, or full name if no comma)
        last = first.split(",")[0].strip() if "," in first else first.split()[-1] if first else ""
        # Remove non-alphanumeric
        last = re.sub(r"[^A-Za-z]", "", last)
    else:
        last = "Unknown"
    yr = str(year)[:4] if year else "nd"
    base = f"{last}{yr}"
    if not base[0].isalpha():
        base = "R" + base
    return base


def _escape_bibtex(text: str) -> str:
    """Escape special BibTeX characters."""
    if not text:
        return ""
    return text.replace("{", "\\{").replace("}", "\\}").replace("&", "\\&")


def export_results_bibtex(
    project_path: str,
    output_path: str | None = None,
    deduplicated: bool = True,
) -> str:
    """Export stored results to a BibTeX file. Returns the output path."""
    results = get_results(project_path, deduplicated=deduplicated)
    if not results:
        return ""

    if output_path is None:
        suffix = "_deduplicated" if deduplicated else ""
        output_path = str(Path(project_path) / "data" / f"search_results{suffix}.bib")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    seen_keys: dict[str, int] = {}
    entries: list[str] = []

    for r in results:
        authors = _parse_authors(r.get("authors_json"))
        year = r.get("year", "")
        doi = r.get("doi", "")
        result_id = r.get("result_id", 0)

        key = _make_citekey(authors, year, doi, result_id)
        count = seen_keys.get(key, 0)
        seen_keys[key] = count + 1
        if count > 0:
            key = f"{key}{chr(ord('a') + count)}"

        author_str = _escape_bibtex(" and ".join(authors))
        title = _escape_bibtex(r.get("title", ""))
        journal = _escape_bibtex(r.get("journal", ""))
        abstract_text = _escape_bibtex(r.get("abstract", ""))

        entry_lines = [f"@article{{{key},"]
        if author_str:
            entry_lines.append(f"  author = {{{author_str}}},")
        if title:
            entry_lines.append(f"  title = {{{{{title}}}}},")
        if journal:
            entry_lines.append(f"  journal = {{{journal}}},")
        if year:
            entry_lines.append(f"  year = {{{year}}},")
        if r.get("volume"):
            entry_lines.append(f"  volume = {{{r['volume']}}},")
        if r.get("issue"):
            entry_lines.append(f"  number = {{{r['issue']}}},")
        if r.get("pages"):
            entry_lines.append(f"  pages = {{{r['pages']}}},")
        if doi:
            entry_lines.append(f"  doi = {{{doi}}},")
        if r.get("pmid"):
            entry_lines.append(f"  pmid = {{{r['pmid']}}},")
        if abstract_text:
            entry_lines.append(f"  abstract = {{{abstract_text}}},")
        entry_lines.append("}")
        entries.append("\n".join(entry_lines))

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n\n".join(entries) + "\n")

    return output_path


def export_results_ris(
    project_path: str,
    output_path: str | None = None,
    deduplicated: bool = True,
) -> str:
    """Export stored results to an RIS file. Returns the output path."""
    results = get_results(project_path, deduplicated=deduplicated)
    if not results:
        return ""

    if output_path is None:
        suffix = "_deduplicated" if deduplicated else ""
        output_path = str(Path(project_path) / "data" / f"search_results{suffix}.ris")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    entries: list[str] = []
    for r in results:
        authors = _parse_authors(r.get("authors_json"))
        lines = ["TY  - JOUR"]
        if r.get("title"):
            lines.append(f"TI  - {r['title']}")
        for a in authors:
            lines.append(f"AU  - {a}")
        if r.get("journal"):
            lines.append(f"JO  - {r['journal']}")
        if r.get("year"):
            lines.append(f"PY  - {r['year']}")
        if r.get("volume"):
            lines.append(f"VL  - {r['volume']}")
        if r.get("issue"):
            lines.append(f"IS  - {r['issue']}")
        if r.get("pages"):
            lines.append(f"SP  - {r['pages']}")
        if r.get("doi"):
            lines.append(f"DO  - {r['doi']}")
        if r.get("pmid"):
            lines.append(f"AN  - {r['pmid']}")
        if r.get("abstract"):
            lines.append(f"AB  - {r['abstract']}")
        lines.append("ER  - ")
        entries.append("\n".join(lines))

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("\n".join(entries) + "\n")

    return output_path


def export_results_csljson(
    project_path: str,
    output_path: str | None = None,
    deduplicated: bool = True,
) -> str:
    """Export stored results to a CSL-JSON file. Returns the output path."""
    results = get_results(project_path, deduplicated=deduplicated)
    if not results:
        return ""

    if output_path is None:
        suffix = "_deduplicated" if deduplicated else ""
        output_path = str(Path(project_path) / "data" / f"search_results{suffix}.json")

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    items: list[dict[str, Any]] = []
    for r in results:
        authors = _parse_authors(r.get("authors_json"))
        year = r.get("year", "")

        csl_authors = []
        for a in authors:
            if "," in a:
                parts = a.split(",", 1)
                csl_authors.append({"family": parts[0].strip(), "given": parts[1].strip()})
            else:
                words = a.strip().split()
                if len(words) >= 2:
                    csl_authors.append({"family": words[-1], "given": " ".join(words[:-1])})
                elif words:
                    csl_authors.append({"family": words[0]})

        item: dict[str, Any] = {
            "type": "article-journal",
            "title": r.get("title", ""),
        }
        if csl_authors:
            item["author"] = csl_authors
        if r.get("journal"):
            item["container-title"] = r["journal"]
        if year:
            if year[:4].isdigit():
                item["issued"] = {"date-parts": [[int(year[:4])]]}
            else:
                item["issued"] = {"literal": year}
        if r.get("volume"):
            item["volume"] = r["volume"]
        if r.get("issue"):
            item["issue"] = r["issue"]
        if r.get("pages"):
            item["page"] = r["pages"]
        if r.get("doi"):
            item["DOI"] = r["doi"]
        if r.get("pmid"):
            item["PMID"] = r["pmid"]
        if r.get("abstract"):
            item["abstract"] = r["abstract"]

        # Use DOI or PMID as id
        item["id"] = r.get("doi") or r.get("pmid") or f"result-{r.get('result_id', '')}"
        items.append(item)

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(items, f, indent=2, ensure_ascii=False)

    return output_path


def export_results_bibliography(
    project_path: str,
    fmt: str = "bibtex",
    output_path: str | None = None,
    deduplicated: bool = True,
) -> str:
    """Export stored results in the specified bibliographic format.

    Args:
        project_path: Absolute path to the project directory.
        fmt: One of ``"bibtex"``, ``"ris"``, ``"csljson"``.
        output_path: Optional custom output file path.
        deduplicated: Whether to use deduplicated results.

    Returns:
        The output file path, or ``""`` if no results.
    """
    exporters = {
        "bibtex": export_results_bibtex,
        "ris": export_results_ris,
        "csljson": export_results_csljson,
    }
    func = exporters.get(fmt)
    if func is None:
        raise ValueError(f"Unsupported format: {fmt!r}. Use one of: {list(exporters)}")
    return func(project_path, output_path=output_path, deduplicated=deduplicated)


# ---------------------------------------------------------------------------
# Attachment management
# ---------------------------------------------------------------------------


def link_file(
    project_path: str,
    result_id: int,
    file_path: str,
    mime_type: str | None = None,
    label: str | None = None,
) -> int:
    """Link a file (PDF, supplementary material, etc.) to a search result.

    Args:
        project_path: Absolute path to the project directory.
        result_id: The result to attach the file to.
        file_path: Absolute or project-relative path to the file.
        mime_type: Optional MIME type (e.g. 'application/pdf').
        label: Optional human-readable label.

    Returns:
        The attachment_id of the new record.
    """
    conn = init_db(project_path)
    try:
        now = datetime.now(UTC).isoformat()
        cursor = conn.execute(
            "INSERT INTO attachments (result_id, file_path, mime_type, label, added_at) "
            "VALUES (?, ?, ?, ?, ?)",
            (result_id, file_path, mime_type, label, now),
        )
        conn.commit()
        return cursor.lastrowid
    finally:
        conn.close()


def get_attachments(
    project_path: str,
    result_id: int,
) -> list[dict[str, Any]]:
    """Return all attachments linked to a result."""
    db = _db_path(project_path)
    if not db.exists():
        return []
    conn = init_db(project_path)
    conn.row_factory = sqlite3.Row
    try:
        rows = conn.execute(
            "SELECT * FROM attachments WHERE result_id = ?", (result_id,)
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Notes management
# ---------------------------------------------------------------------------


def add_note(
    project_path: str,
    result_id: int,
    content: str,
    note_type: str = "general",
    tags: list[str] | None = None,
) -> int:
    """Add a note to a search result.

    Args:
        project_path: Absolute path to the project directory.
        result_id: The result to annotate.
        content: The note text (Markdown supported).
        note_type: Category — ``'general'``, ``'critique'``, ``'summary'``,
            ``'extraction'``, ``'methodology'``.
        tags: Optional list of freeform tags.

    Returns:
        The note_id of the new record.
    """
    conn = init_db(project_path)
    try:
        now = datetime.now(UTC).isoformat()
        cursor = conn.execute(
            "INSERT INTO notes (result_id, note_type, content, tags_json, created_at, updated_at) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            (result_id, note_type, content, json.dumps(tags) if tags else None, now, now),
        )
        conn.commit()
        return cursor.lastrowid
    finally:
        conn.close()


def update_note(
    project_path: str,
    note_id: int,
    content: str | None = None,
    note_type: str | None = None,
    tags: list[str] | None = None,
) -> bool:
    """Update an existing note. Returns True if the note was found and updated."""
    db = _db_path(project_path)
    if not db.exists():
        return False
    conn = init_db(project_path)
    try:
        sets: list[str] = []
        params: list[Any] = []
        if content is not None:
            sets.append("content = ?")
            params.append(content)
        if note_type is not None:
            sets.append("note_type = ?")
            params.append(note_type)
        if tags is not None:
            sets.append("tags_json = ?")
            params.append(json.dumps(tags))
        if not sets:
            return False
        sets.append("updated_at = ?")
        params.append(datetime.now(UTC).isoformat())
        params.append(note_id)
        cursor = conn.execute(f"UPDATE notes SET {', '.join(sets)} WHERE note_id = ?", params)
        conn.commit()
        return cursor.rowcount > 0
    finally:
        conn.close()


def get_notes(
    project_path: str,
    result_id: int | None = None,
    note_type: str | None = None,
) -> list[dict[str, Any]]:
    """Query notes, optionally filtering by result and/or type."""
    db = _db_path(project_path)
    if not db.exists():
        return []
    conn = init_db(project_path)
    conn.row_factory = sqlite3.Row
    try:
        where: list[str] = []
        params: list[Any] = []
        if result_id is not None:
            where.append("result_id = ?")
            params.append(result_id)
        if note_type is not None:
            where.append("note_type = ?")
            params.append(note_type)
        sql = "SELECT * FROM notes"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY created_at DESC"
        rows = conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def search_notes(
    project_path: str,
    query: str,
) -> list[dict[str, Any]]:
    """Full-text search across note content and tags."""
    db = _db_path(project_path)
    if not db.exists():
        return []
    conn = init_db(project_path)
    conn.row_factory = sqlite3.Row
    try:
        pattern = f"%{query}%"
        rows = conn.execute(
            "SELECT n.*, r.title AS result_title, r.doi, r.pmid "
            "FROM notes n JOIN results r ON n.result_id = r.result_id "
            "WHERE n.content LIKE ? OR n.tags_json LIKE ? "
            "ORDER BY n.updated_at DESC",
            (pattern, pattern),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# Annotation storage (extracted from PDFs)
# ---------------------------------------------------------------------------


def store_annotations(
    project_path: str,
    attachment_id: int,
    result_id: int,
    annotations: list[dict[str, Any]],
) -> int:
    """Bulk-store annotations extracted from a PDF attachment.

    Each annotation dict should have optional keys:
    ``page``, ``annotation_type``, ``color``, ``content``, ``comment``.

    Returns:
        Number of annotations stored.
    """
    conn = init_db(project_path)
    try:
        now = datetime.now(UTC).isoformat()
        count = 0
        for a in annotations:
            conn.execute(
                "INSERT INTO annotations "
                "(attachment_id, result_id, page, annotation_type, color, "
                "content, comment, extracted_at) "
                "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (
                    attachment_id,
                    result_id,
                    a.get("page"),
                    a.get("annotation_type", "highlight"),
                    a.get("color"),
                    a.get("content", ""),
                    a.get("comment", ""),
                    now,
                ),
            )
            count += 1
        conn.commit()
        return count
    finally:
        conn.close()


def get_annotations(
    project_path: str,
    result_id: int | None = None,
    attachment_id: int | None = None,
    color: str | None = None,
) -> list[dict[str, Any]]:
    """Query stored annotations with optional filters."""
    db = _db_path(project_path)
    if not db.exists():
        return []
    conn = init_db(project_path)
    conn.row_factory = sqlite3.Row
    try:
        where: list[str] = []
        params: list[Any] = []
        if result_id is not None:
            where.append("result_id = ?")
            params.append(result_id)
        if attachment_id is not None:
            where.append("attachment_id = ?")
            params.append(attachment_id)
        if color is not None:
            where.append("color = ?")
            params.append(color)
        sql = "SELECT * FROM annotations"
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY page, annotation_id"
        rows = conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def search_annotations(
    project_path: str,
    query: str,
) -> list[dict[str, Any]]:
    """Search annotation content and comments."""
    db = _db_path(project_path)
    if not db.exists():
        return []
    conn = init_db(project_path)
    conn.row_factory = sqlite3.Row
    try:
        pattern = f"%{query}%"
        rows = conn.execute(
            "SELECT a.*, r.title AS result_title, r.doi "
            "FROM annotations a JOIN results r ON a.result_id = r.result_id "
            "WHERE a.content LIKE ? OR a.comment LIKE ? "
            "ORDER BY a.result_id, a.page",
            (pattern, pattern),
        ).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


# ---------------------------------------------------------------------------
# BibTeX / RIS import
# ---------------------------------------------------------------------------


def _parse_bibtex_entries(text: str) -> list[dict[str, Any]]:
    """Minimal BibTeX parser — handles @article{key, field={value}} entries."""
    entries: list[dict[str, Any]] = []
    # Split on @ followed by an entry type
    raw_entries = re.split(r"(?=@\w+\{)", text.strip())
    for raw in raw_entries:
        raw = raw.strip()
        if not raw:
            continue
        # Match entry type and key
        m = re.match(r"@(\w+)\{([^,]*),?(.*)\}", raw, re.DOTALL)
        if not m:
            continue
        entry: dict[str, Any] = {"_type": m.group(1).lower(), "_key": m.group(2).strip()}
        body = m.group(3)
        # Extract field = {value} or field = "value"
        for fm in re.finditer(r"(\w+)\s*=\s*(?:\{((?:[^{}]|\{[^{}]*\})*)\}|\"([^\"]*)\")", body):
            field = fm.group(1).lower()
            value = (fm.group(2) or fm.group(3) or "").strip()
            entry[field] = value
        entries.append(entry)
    return entries


def import_bibtex(
    project_path: str,
    bibtex_text: str | None = None,
    file_path: str | None = None,
    source: str = "bibtex_import",
) -> int:
    """Import references from BibTeX text or file into the result store.

    Returns:
        The search_id under which the imported items are stored.
    """
    if file_path:
        bibtex_text = Path(file_path).read_text(encoding="utf-8")
    if not bibtex_text:
        return 0

    entries = _parse_bibtex_entries(bibtex_text)
    if not entries:
        return 0

    results: list[dict[str, Any]] = []
    for e in entries:
        authors_raw = e.get("author", "")
        authors = [a.strip() for a in authors_raw.split(" and ")] if authors_raw else []
        results.append(
            {
                "doi": e.get("doi", ""),
                "pmid": e.get("pmid", ""),
                "title": e.get("title", ""),
                "authors": authors,
                "journal": e.get("journal", ""),
                "year": e.get("year", ""),
                "volume": e.get("volume", ""),
                "issue": e.get("number", ""),
                "pages": e.get("pages", ""),
                "abstract": e.get("abstract", ""),
                "_bibtex_key": e.get("_key", ""),
                "_bibtex_type": e.get("_type", ""),
            }
        )

    return store_results(
        project_path,
        source,
        f"BibTeX import ({len(results)} entries)",
        results,
        total_count=len(results),
        parameters={"format": "bibtex", "file_path": file_path},
    )


def _parse_ris_entries(text: str) -> list[dict[str, Any]]:
    """Minimal RIS parser."""
    entries: list[dict[str, Any]] = []
    current: dict[str, Any] = {}
    current_authors: list[str] = []

    for line in text.splitlines():
        line = line.rstrip()
        if not line:
            continue
        # RIS lines: XX  - value
        m = re.match(r"^([A-Z][A-Z0-9])\s{2}-\s?(.*)", line)
        if not m:
            continue
        tag, value = m.group(1), m.group(2).strip()
        if tag == "TY":
            current = {"_type": value}
            current_authors = []
        elif tag == "ER":
            if current:
                current["authors"] = current_authors
                entries.append(current)
            current = {}
            current_authors = []
        elif tag == "AU":
            current_authors.append(value)
        elif tag == "TI" or tag == "T1":
            current["title"] = value
        elif tag == "JO" or tag == "JF" or tag == "T2":
            current.setdefault("journal", value)
        elif tag == "PY" or tag == "Y1":
            current["year"] = value[:4] if len(value) >= 4 else value
        elif tag == "VL":
            current["volume"] = value
        elif tag == "IS":
            current["issue"] = value
        elif tag == "SP":
            current["pages"] = value
        elif tag == "EP":
            sp = current.get("pages", "")
            if sp:
                current["pages"] = f"{sp}-{value}"
            else:
                current["pages"] = value
        elif tag == "DO":
            current["doi"] = value
        elif tag == "AN":
            current["pmid"] = value
        elif tag == "AB":
            current["abstract"] = value
    # If file doesn't end with ER
    if current:
        current["authors"] = current_authors
        entries.append(current)
    return entries


def import_ris(
    project_path: str,
    ris_text: str | None = None,
    file_path: str | None = None,
    source: str = "ris_import",
) -> int:
    """Import references from RIS text or file into the result store.

    Returns:
        The search_id under which the imported items are stored.
    """
    if file_path:
        ris_text = Path(file_path).read_text(encoding="utf-8")
    if not ris_text:
        return 0

    entries = _parse_ris_entries(ris_text)
    if not entries:
        return 0

    results: list[dict[str, Any]] = []
    for e in entries:
        results.append(
            {
                "doi": e.get("doi", ""),
                "pmid": e.get("pmid", ""),
                "title": e.get("title", ""),
                "authors": e.get("authors", []),
                "journal": e.get("journal", ""),
                "year": e.get("year", ""),
                "volume": e.get("volume", ""),
                "issue": e.get("issue", ""),
                "pages": e.get("pages", ""),
                "abstract": e.get("abstract", ""),
            }
        )

    return store_results(
        project_path,
        source,
        f"RIS import ({len(results)} entries)",
        results,
        total_count=len(results),
        parameters={"format": "ris", "file_path": file_path},
    )


# ---------------------------------------------------------------------------
# Reading list / status tracking
# ---------------------------------------------------------------------------


def get_reading_list(
    project_path: str,
    note_type: str | None = None,
    has_attachments: bool | None = None,
    has_notes: bool | None = None,
) -> list[dict[str, Any]]:
    """Return a reading-list view joining results with attachment/note counts.

    Args:
        project_path: Absolute path to the project directory.
        note_type: Optional filter — only include results that have a note of this type.
        has_attachments: If True, only results with linked files.
            If False, only results without.
        has_notes: If True, only results with notes. If False, only without.

    Returns:
        List of dicts with result fields plus ``attachment_count``,
        ``note_count``, ``annotation_count``.
    """
    db = _db_path(project_path)
    if not db.exists():
        return []
    conn = init_db(project_path)
    conn.row_factory = sqlite3.Row
    try:
        sql = (
            "SELECT r.result_id, r.doi, r.pmid, r.title, r.authors_json, "
            "r.journal, r.year, r.source, "
            "COALESCE(att.cnt, 0) AS attachment_count, "
            "COALESCE(nt.cnt, 0) AS note_count, "
            "COALESCE(ann.cnt, 0) AS annotation_count "
            "FROM results r "
            "LEFT JOIN (SELECT result_id, COUNT(*) AS cnt FROM attachments GROUP BY result_id) att "
            "ON r.result_id = att.result_id "
            "LEFT JOIN (SELECT result_id, COUNT(*) AS cnt FROM notes GROUP BY result_id) nt "
            "ON r.result_id = nt.result_id "
            "LEFT JOIN (SELECT result_id, COUNT(*) AS cnt FROM annotations GROUP BY result_id) ann "
            "ON r.result_id = ann.result_id "
        )
        where: list[str] = []
        params: list[Any] = []

        if note_type is not None:
            where.append("r.result_id IN (SELECT result_id FROM notes WHERE note_type = ?)")
            params.append(note_type)
        if has_attachments is True:
            where.append("COALESCE(att.cnt, 0) > 0")
        elif has_attachments is False:
            where.append("COALESCE(att.cnt, 0) = 0")
        if has_notes is True:
            where.append("COALESCE(nt.cnt, 0) > 0")
        elif has_notes is False:
            where.append("COALESCE(nt.cnt, 0) = 0")

        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY r.result_id"

        rows = conn.execute(sql, params).fetchall()
        return [dict(r) for r in rows]
    finally:
        conn.close()


def register_result_store_tools(mcp_instance: Any) -> None:
    """Register search result query/export tools on an MCP server instance.

    Call this at module level in each search server to add
    get_stored_results and export_stored_results tools.
    """

    @mcp_instance.tool()
    async def get_stored_results(
        project_path: str,
        source: str | None = None,
        deduplicated: bool = False,
    ) -> dict[str, Any]:
        """Query previously stored search results from the project database.

        Args:
            project_path: Absolute path to the project directory.
            source: Optional filter by source (e.g., 'pubmed', 'openalex').
            deduplicated: If True, return deduplicated results across all sources.

        Returns:
            Dictionary with total counts, result list, and search history.
        """
        results = get_results(project_path, source=source, deduplicated=deduplicated)
        searches = get_searches(project_path)
        return {
            "total_results": len(results),
            "total_searches": len(searches),
            "results": results,
            "searches": searches,
        }

    @mcp_instance.tool()
    async def export_stored_results(
        project_path: str,
        output_path: str | None = None,
        deduplicated: bool = False,
    ) -> dict[str, str]:
        """Export stored search results to an Excel workbook with clickable DOI/PMID links.

        Args:
            project_path: Absolute path to the project directory.
            output_path: Optional custom output file path. Defaults to
                {project_path}/data/search_results.xlsx.
            deduplicated: If True, export deduplicated results.

        Returns:
            Dictionary with export status and file path.
        """
        path = export_results_excel(project_path, output_path, deduplicated)
        if path:
            return {"status": "exported", "file": path}
        return {"status": "no_results", "file": ""}

    @mcp_instance.tool()
    async def export_stored_bibliography(
        project_path: str,
        format: str = "bibtex",
        output_path: str | None = None,
        deduplicated: bool = True,
    ) -> dict[str, str]:
        """Export stored search results to a bibliographic format (BibTeX, RIS, or CSL-JSON).

        Args:
            project_path: Absolute path to the project directory.
            format: Export format — 'bibtex' (default), 'ris', or 'csljson'.
            output_path: Optional custom output file path.
            deduplicated: If True (default), export deduplicated results.

        Returns:
            Dictionary with export status and file path.
        """
        try:
            path = export_results_bibliography(
                project_path,
                fmt=format,
                output_path=output_path,
                deduplicated=deduplicated,
            )
        except ValueError as exc:
            return {"status": "error", "file": "", "message": str(exc)}
        if path:
            return {"status": "exported", "file": path}
        return {"status": "no_results", "file": ""}
